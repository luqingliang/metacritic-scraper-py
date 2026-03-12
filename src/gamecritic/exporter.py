from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _fetch_rows(conn: sqlite3.Connection, query: str, params: tuple[Any, ...]) -> list[dict]:
    cursor = conn.execute(query, params)
    columns = [desc[0] for desc in cursor.description]
    rows = []
    for values in cursor.fetchall():
        rows.append(dict(zip(columns, values)))
    return rows


def _table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    cursor = conn.execute(f"PRAGMA table_info({table_name})")
    return {str(row[1]) for row in cursor.fetchall()}


def _write_sheet(ws: Worksheet, rows: list[dict], columns: list[str]) -> int:
    ws.append(columns)
    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in rows:
        ws.append([row.get(col) for col in columns])

    # Autosize to keep the file readable for manual checks.
    for idx, col in enumerate(columns, start=1):
        max_len = len(str(col))
        for row in rows:
            value = row.get(col)
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 2, 80)

    ws.freeze_panes = "A2"
    return len(rows)


def export_sqlite_to_excel(
    *,
    db_path: str | Path,
    output_path: str | Path,
) -> dict[str, int]:
    db_path = Path(db_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not db_path.exists():
        raise FileNotFoundError(f"database not found: {db_path}")

    games_columns = [
        "slug",
        "game_id",
        "title",
        "platform",
        "release_date",
        "premiere_year",
        "rating",
        "critic_score",
        "critic_review_count",
        "user_score",
        "user_review_count",
        "cover_url",
        "scraped_at",
    ]

    critic_columns = [
        "slug",
        "review_key",
        "score",
        "review_date",
        "author",
        "publication_name",
        "source_url",
        "quote",
        "scraped_at",
    ]

    user_columns = [
        "review_id",
        "slug",
        "author",
        "score",
        "review_date",
        "spoiler",
        "quote",
        "scraped_at",
    ]

    params: tuple[Any, ...] = tuple()

    with sqlite3.connect(db_path) as conn:
        games_columns = [col for col in games_columns if col in _table_columns(conn, "games")]
        critic_columns = [col for col in critic_columns if col in _table_columns(conn, "critic_reviews")]
        user_columns = [col for col in user_columns if col in _table_columns(conn, "user_reviews")]

        games_query = f"SELECT {', '.join(games_columns)} FROM games"
        critic_query = f"SELECT {', '.join(critic_columns)} FROM critic_reviews"
        user_query = f"SELECT {', '.join(user_columns)} FROM user_reviews"

        games_rows = _fetch_rows(conn, games_query, params)
        critic_rows = _fetch_rows(conn, critic_query, params)
        user_rows = _fetch_rows(conn, user_query, params)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    summary_rows = [
        ("generated_at", _utc_now_iso()),
        ("database", str(db_path)),
        ("games_rows", len(games_rows)),
        ("critic_reviews_rows", len(critic_rows)),
        ("user_reviews_rows", len(user_rows)),
    ]
    summary_ws.append(["field", "value"])
    for cell in summary_ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    for row in summary_rows:
        summary_ws.append(list(row))
    summary_ws.column_dimensions["A"].width = 26
    summary_ws.column_dimensions["B"].width = 80
    summary_ws.freeze_panes = "A2"

    games_ws = workbook.create_sheet("Games")
    critic_ws = workbook.create_sheet("CriticReviews")
    user_ws = workbook.create_sheet("UserReviews")

    counts = {
        "games_rows": _write_sheet(games_ws, games_rows, games_columns),
        "critic_reviews_rows": _write_sheet(critic_ws, critic_rows, critic_columns),
        "user_reviews_rows": _write_sheet(user_ws, user_rows, user_columns),
    }

    workbook.save(output_path)
    return counts
